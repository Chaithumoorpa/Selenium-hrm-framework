"""
Enterprise Selenium HRM Framework — Excel Utilities
----------------------------------------------------
Reads test data from Excel (.xlsx) files using openpyxl.
Supports sheet-based data separation by module.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from utilities.logger import get_logger

logger = get_logger(__name__)


class ExcelUtils:
    """
    Excel test data reader using openpyxl.

    Convention:
        - Each sheet = one test module (Login, Employee, Leave, etc.)
        - Row 1 = column headers
        - Rows 2+ = test data records

    Usage:
        excel = ExcelUtils("testdata/test_data.xlsx")
        users = excel.get_sheet_data("Login")
        # → [{"username": "Admin", "password": "admin123", "expected": "Dashboard"}]
    """

    def __init__(self, file_path: str) -> None:
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Test data file not found: {self.file_path}")
        self._workbook: Optional[openpyxl.Workbook] = None
        logger.info("excel_utils_init", file=str(self.file_path))

    @property
    def workbook(self) -> openpyxl.Workbook:
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(
                str(self.file_path), read_only=True, data_only=True
            )
            logger.debug("excel_workbook_loaded", sheets=self._workbook.sheetnames)
        return self._workbook

    def get_sheet_names(self) -> List[str]:
        """Return list of all sheet names in the workbook."""
        return self.workbook.sheetnames

    def get_sheet_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Read all rows from a sheet as list of dicts.

        Args:
            sheet_name: Name of the sheet (e.g., "Login", "Employee")

        Returns:
            List of row dicts with column headers as keys

        Raises:
            KeyError: If sheet does not exist
        """
        if sheet_name not in self.workbook.sheetnames:
            raise KeyError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {self.workbook.sheetnames}"
            )

        sheet: Worksheet = self.workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            logger.warning("excel_empty_sheet", sheet=sheet_name)
            return []

        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if all(cell is None for cell in row):
                continue  # Skip blank rows
            row_dict = {
                headers[i]: str(cell).strip() if cell is not None else ""
                for i, cell in enumerate(row)
                if i < len(headers)
            }
            data.append(row_dict)

        logger.info("excel_data_read", sheet=sheet_name, rows=len(data))
        return data

    def get_row_by_id(
        self, sheet_name: str, id_column: str, id_value: str
    ) -> Optional[Dict[str, Any]]:
        """
        Get a single row matching an ID value.

        Args:
            sheet_name: Sheet to search
            id_column:  Column name to match on
            id_value:   Value to find

        Returns:
            Row dict or None if not found
        """
        data = self.get_sheet_data(sheet_name)
        for row in data:
            if row.get(id_column) == id_value:
                return row
        logger.warning(
            "excel_row_not_found",
            sheet=sheet_name,
            column=id_column,
            value=id_value,
        )
        return None

    def get_column_values(self, sheet_name: str, column_name: str) -> List[str]:
        """Get all values from a specific column."""
        return [
            row[column_name]
            for row in self.get_sheet_data(sheet_name)
            if column_name in row
        ]

    def close(self) -> None:
        """Close workbook and release file handle."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            logger.debug("excel_workbook_closed")

    def __enter__(self) -> "ExcelUtils":
        return self

    def __exit__(self, *args) -> None:
        self.close()


def load_json_data(file_path: str) -> Any:
    """Load JSON test data file."""
    import json

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"JSON data file not found: {path}")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    logger.debug("json_data_loaded", file=str(path))
    return data
